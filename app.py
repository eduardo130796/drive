import requests
import pandas as pd
import openpyxl
from datetime import datetime
import io
import streamlit as st
import time


## Configuração da página
st.set_page_config(page_title="Gestão de Planilhas", page_icon="📊", layout="wide")

# Estilo personalizado
st.markdown(
    """
    <style>
    body {
        font-family: 'Arial', sans-serif;
    }
    .stButton>button {
        background-color: #f0f5f0;
        color: black;
        font-size: 16px;
        border-radius: 10px;
        padding: 10px 20px;
    }
    .stDownloadButton>button {
        background-color: #008CBA;
        color: white;
        font-size: 16px;
        border-radius: 10px;
        padding: 10px 20px;
    }
    .stFileUploader {
        border: 2px dashed #ddd;
        padding: 10px;
        border-radius: 10px;
        text-align: center;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# Suas credenciais OAuth 2.0
CLIENT_ID = "652191149879-n4l39h8quk3rfg4lmb2ijeb4pbm347af.apps.googleusercontent.com"
CLIENT_SECRET = "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"
REFRESH_TOKEN = "1//04kwvkYH368F3CgYIARAAGAQSNwF-L9Iroi9uVHdeYlwhisoPmabHUykgNBzeCvl_SgTpewicuRPhRd6FbCDLcJd4E3_h0JqSHEo"  # Use o refresh_token que você obteve na primeira autenticação

# ID da pasta específica
PASTA_ID = "1MP1wiSpSBx_pNTt3PTo3ayokZFxm_niR"  # Coloque o ID da sua pasta aqui

# Função para renovar o token usando o refresh_token
def renovar_token(refresh_token):
    url = "https://oauth2.googleapis.com/token"
    data = {
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
        'refresh_token': refresh_token,
        'grant_type': 'refresh_token'
    }
    
    response = requests.post(url, data=data)
    if response.status_code == 200:
        token_info = response.json()
        access_token = token_info.get('access_token')
        return access_token
    else:
        return None
# Função para obter o token de acesso
def get_access_token():
    global ACCESS_TOKEN
    if not hasattr(get_access_token, "expires_at") or time.time() > get_access_token.expires_at:
        # Se o token tiver expirado ou não estiver definido, renove-o
        ACCESS_TOKEN = renovar_token(REFRESH_TOKEN)
        get_access_token.expires_at = time.time() + 3600  # Defina o tempo de expiração como 1 hora (3600 segundos)
    return ACCESS_TOKEN

# Função para listar arquivos dentro da pasta
def listar_arquivos(pasta_id):
    url = "https://www.googleapis.com/drive/v3/files"
    params = {
        "q": f"'{pasta_id}' in parents and trashed=false",  # Busca apenas arquivos na pasta especificada
        "fields": "files(id, name)"
    }
    headers = {"Authorization": f"Bearer {get_access_token()}"}

    response = requests.get(url, params=params, headers=headers)
    arquivos = response.json().get("files", [])
    
    if arquivos:
        print(f"📂 Arquivos na pasta {pasta_id}:")
        for arquivo in arquivos:
            print(f"📄 {arquivo['name']} (ID: {arquivo['id']})")
        return arquivos
    else:
        print("❌ Nenhum arquivo encontrado na pasta.")
        return []

# Função para baixar o arquivo do Google Drive
def baixar_arquivo_drive(file_id):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    headers = {"Authorization": f"Bearer {get_access_token()}"}
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        return io.BytesIO(response.content)  # Retorna o arquivo como BytesIO
    else:
        raise Exception(f"Erro ao baixar o arquivo: {response.status_code}")
    



# Função para atualizar a planilha no Google Drive
def atualizar_planilha_drive(file_id, arquivo_processado):
    url = f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media"
    headers = {
        "Authorization": f"Bearer {get_access_token()}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # Definindo o tipo do arquivo
    }
    
    # O conteúdo do arquivo processado será enviado
    response = requests.patch(url, headers=headers, data=arquivo_processado)

    if response.status_code == 200:
        print(f"✅ Arquivo {file_id} atualizado com sucesso!")
        return True
    else:
        print(f"❌ Erro ao atualizar o arquivo: {response.status_code}")
        print(response.text)  # Adiciona mais detalhes sobre o erro
        return False

# Função para processar a planilha e registrar alterações no log
def processar_planilhas(arquivo_base, arquivo_atualizacao, nome_arquivo_base):
    # Reabre a planilha para garantir que todas as abas sejam recarregadas
    wb = openpyxl.load_workbook(arquivo_base)
    ws = wb.active  # Aba principal

    # Verifica se a aba "Log de Alterações" existe, se não, cria a aba
    if "Log de Alterações" not in wb.sheetnames:
        ws_log = wb.create_sheet("Log de Alterações")
        # Adiciona o cabeçalho na aba de log
        ws_log.append(["Tipo", "Nota de Empenho", "Campo", "Valor Antigo", "Valor Novo", "Data da Alteração"])
    else:
        ws_log = wb["Log de Alterações"]

    # Processamento da atualização
    df_atualizacao = pd.read_excel(arquivo_atualizacao, skiprows=2)

    mapa_valor_empenhado = {
        str(row["Número da Nota de Empenho"]).strip()[-8:]: row["Saldo - R$ (Item Informação)"]
        for _, row in df_atualizacao.iterrows()
    }

    # Processamento das notas de empenho
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[4]
        valor_empenhado_cell = row[5]
        nota_empenho = str(nota_empenho_cell.value).strip()

        if nota_empenho in mapa_valor_empenhado:
            novo_valor = mapa_valor_empenhado[nota_empenho]
            if str(valor_empenhado_cell.value).strip() != str(novo_valor).strip():
                # Registro no log
                ws_log.append(["Empenho", nota_empenho, "Valor Empenhado", valor_empenhado_cell.value, novo_valor, 
                               pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
                valor_empenhado_cell.value = novo_valor

    # Atualização de pagamentos
    df_atualizacao = df_atualizacao[~df_atualizacao.apply(lambda row: row.astype(str).str.contains('Total').any(), axis=1)]
    df_atualizacao.ffill(inplace=True)

    meses = {"jan": 12, "fev": 13, "mar": 14, "abr": 15, "mai": 16, "jun": 17,
             "jul": 18, "ago": 19, "set": 20, "out": 21, "nov": 22, "dez": 23}
    meses_ingles_para_portugues = {'jan': 'jan', 'feb': 'fev', 'mar': 'mar', 'apr': 'abr', 'may': 'mai',
                                   'jun': 'jun', 'jul': 'jul', 'aug': 'ago', 'sep': 'set', 'oct': 'out',
                                   'nov': 'nov', 'dec': 'dez'}

    pagamentos_por_nota = {}
    for _, row in df_atualizacao.iterrows():
        nota_empenho = str(row["Número da Nota de Empenho"]).strip()[-8:]
        data_pagamento = row["Métrica"]
        valor_pago = row["Unnamed: 13"]

        if nota_empenho not in pagamentos_por_nota:
            pagamentos_por_nota[nota_empenho] = {mes: [] for mes in meses}

        data_pagamento = pd.to_datetime(data_pagamento, errors='coerce', dayfirst=True)
        if pd.notna(data_pagamento):
            mes_pagamento = data_pagamento.strftime('%b').lower()
            mes_pagamento_portugues = meses_ingles_para_portugues.get(mes_pagamento)
            if mes_pagamento_portugues in meses:
                pagamentos_por_nota[nota_empenho][mes_pagamento_portugues].append(float(valor_pago))

    # Atualiza as células de pagamento e registra no log se houver alteração
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[4]
        status_cell = row[10]
        nota_empenho = str(nota_empenho_cell.value).strip()

        if nota_empenho in pagamentos_por_nota:
            for mes, coluna_mes in meses.items():
                valores_novos = pagamentos_por_nota[nota_empenho].get(mes, [])

                if valores_novos:
                    valor_pago_cell = row[coluna_mes]

                    if isinstance(valores_novos, (int, float)):
                        valores_novos = [valores_novos]

                    status_texto = str(status_cell.value).strip().lower() if status_cell.value else ""

                    status_aceitos = [
                        "não pediu, mas pode solicitar.",
                        "solicitado - em análise",
                        "Não solicitou, mas pode pedir"
                    ]

                    if any(status_texto.startswith(opcao.lower()) for opcao in status_aceitos):
                        soma_valores = "+".join(str(v).replace(",", ".") for v in valores_novos)
                        nova_formula = f"=({soma_valores})+(({soma_valores})*AE6)"
                    else:
                        soma_valores = "+".join(str(v).replace(",", ".") for v in valores_novos)
                        nova_formula = f"={soma_valores}" if len(valores_novos) > 1 else f"={valores_novos[0]}"

                    if str(valor_pago_cell.value).strip() != nova_formula:
                        # Registro no log
                        ws_log.append(["Pagamento", nota_empenho, f"Pagamento {mes}", valor_pago_cell.value, nova_formula, 
                                       pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
                        valor_pago_cell.value = nova_formula

    # Atualiza o cabeçalho com a data e hora da última atualização
    data_hora_atualizacao = datetime.now().strftime("Última atualização: %d/%m/%Y às %H:%M")
    ws["A1"] = data_hora_atualizacao

    # Salva a planilha no buffer para retorno
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer, nome_arquivo_base  # Retorna a planilha atualizada e o nome do arquivo

# Interface Streamlit
#st.set_page_config(page_title="Processamento de Planilhas", page_icon="📊", layout="centered")
# Interface
st.title("📂 Processador de Planilhas no Google Drive")
st.markdown("Melhore sua produtividade automatizando atualizações de planilhas com apenas um clique! 🚀")
st.html("</br></br></br>")
# Gerenciar token
token_valido = False
ACCESS_TOKEN = renovar_token(REFRESH_TOKEN)
if ACCESS_TOKEN:
    token_valido = True
else:
    st.error("Erro ao renovar o token. Insira um novo refresh token para continuar.")
    REFRESH_TOKEN = st.text_input("🔑 Insira um novo Refresh Token", type="password")
    if REFRESH_TOKEN:
        ACCESS_TOKEN = renovar_token(REFRESH_TOKEN)
        if ACCESS_TOKEN:
            st.success("✅ Token atualizado com sucesso!")
            token_valido = True
        else:
            st.error("❌ Falha ao atualizar o token. Verifique o refresh token inserido.")

# Se o token for válido, continuar execução
def get_access_token():
    global ACCESS_TOKEN
    if not hasattr(get_access_token, "expires_at") or time.time() > get_access_token.expires_at:
        ACCESS_TOKEN = renovar_token(REFRESH_TOKEN)
        if ACCESS_TOKEN:
            get_access_token.expires_at = time.time() + 3600  # Define tempo de expiração
    return ACCESS_TOKEN

if token_valido:
    
    col1, col2 = st.columns(2)
    arquivos = listar_arquivos(PASTA_ID)
    with col1:
        if arquivos:
            nomes_arquivos = [arquivo['name'] for arquivo in arquivos]  # Certifique-se de que arquivos é uma lista
            planilhas_selecionadas = st.multiselect("📑 Selecione as planilhas com os objetos", nomes_arquivos)
    
    with col2:
        uploaded_file_atualizacao = st.file_uploader("📤 Selecione a planilha de Notas de Empenho", type=["xlsx"])

    
    if planilhas_selecionadas and uploaded_file_atualizacao:
        if st.button("Iniciar"):
            with st.spinner("🔄 Processando suas planilhas... Isso pode levar alguns segundos."):
                progress_bar = st.progress(0)
                for i, nome_arquivo_base in enumerate(planilhas_selecionadas):
                    arquivo_base_id = next(arquivo['id'] for arquivo in arquivos if arquivo['name'] == nome_arquivo_base)
                    arquivo_base = baixar_arquivo_drive(arquivo_base_id)
                    
                    buffer_final, nome_arquivo = processar_planilhas(arquivo_base, uploaded_file_atualizacao, nome_arquivo_base)
                    
                    if atualizar_planilha_drive(arquivo_base_id, buffer_final):
                        st.success(f"✅ {nome_arquivo} atualizado com sucesso no Google Drive!")

                    progress_bar.progress((i + 1) / len(planilhas_selecionadas))    
                    
                    st.download_button(
                        label=f"📥 Baixar Planilha Finalizada: {nome_arquivo}",
                        data=buffer_final,
                        file_name=nome_arquivo.replace(".xlsx", "_atualizada.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Rodapé fixo com largura total
rodape = """
    <style>
        .footer {
            position: fixed;
            left: 0;
            bottom: 0;
            width: 100%;
            background-color: #f8f9fa;
            text-align: center;
            padding: 10px;
            font-size: 14px;
            color: #6c757d;
            border-top: 1px solid #dee2e6;
            z-index: 100;
        }
    </style>
    <div class="footer">
        Desenvolvido por <strong>Eduardo Júnior</strong> | 2025
    </div>
"""

# Exibir o rodapé na interface
st.markdown(rodape, unsafe_allow_html=True)
